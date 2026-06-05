# Generates docs/FILE_STRUCTURE.xlsx — an English Excel sheet documenting every
# file in the PropertyPulse monorepo (Area | Layer | File | What it does).

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (Area, Layer, File path, What it does)
ROWS = [
    # --- Root ---
    ("Root", "Config", "package.json", "Monorepo definition: npm workspaces (apps/*, packages/*) and run scripts (dev, build, lint, typecheck)."),
    ("Root", "Config", "package-lock.json", "Locks dependency versions across the whole monorepo."),
    ("Root", "Config", "tsconfig.base.json", "Base TypeScript settings every package/app extends (strict, ES2022)."),
    ("Root", "Docs", "README.md", "Project overview and how to run it."),

    # --- packages/shared-types ---
    ("packages/shared-types", "Domain Types", "src/property.types.ts", "Property types: Property, PropertyType, Address, GeoLocation, ListingStatus."),
    ("packages/shared-types", "Domain Types", "src/analysis.types.ts", "Financial analysis types: FinancialAssumptions, InvestmentMetrics, RiskAssessment, RiskFactor."),
    ("packages/shared-types", "Domain Types", "src/report.types.ts", "Investment report types: InvestmentReport, MarketTrendPoint, NeighborhoodInsight."),
    ("packages/shared-types", "Domain Types", "src/user.types.ts", "User types: UserProfile, UserRole, WatchlistItem."),
    ("packages/shared-types", "Domain Types", "src/api.types.ts", "API envelopes: ApiResponse<T>, ApiError, Paginated<T>, ChatMessage."),
    ("packages/shared-types", "Barrel", "src/index.ts", "Re-exports all shared domain types from one entry point."),

    # --- packages/shared-utils ---
    ("packages/shared-utils", "Pure Functions", "src/financial.ts", "Core real-estate math: mortgage payment, rental yields, cap rate, computeInvestmentMetrics()."),
    ("packages/shared-utils", "Pure Functions", "src/formatters.ts", "Display formatting: currency, percent, area, date."),
    ("packages/shared-utils", "Constants", "src/constants.ts", "Defaults/constants: DEFAULT_ASSUMPTIONS, RISK_COLORS, RECOMMENDATION_LABELS."),
    ("packages/shared-utils", "Barrel", "src/index.ts", "Re-exports all shared utilities."),

    # --- packages/ai-core ---
    ("packages/ai-core", "LLM Contracts", "src/llm/types.ts", "LLM types: MessageRole, LlmMessage, GenerationOptions, LlmResponse, TokenUsage."),
    ("packages/ai-core", "LLM Contracts", "src/llm/LlmClient.ts", "LlmClient interface (generate/generateJSON/stream/embed) — implemented by the server's geminiClient."),
    ("packages/ai-core", "RAG Contracts", "src/rag/types.ts", "RAG types: Document, Chunk, ScoredChunk, EmbeddingVector, RetrievalResult."),
    ("packages/ai-core", "RAG Contracts", "src/rag/Retriever.ts", "Retriever interface (embed/search/ingest) — implemented over Supabase pgvector (ragRetriever)."),
    ("packages/ai-core", "Agent Contracts", "src/agents/types.ts", "Agent types: AgentContext, AgentResult, ToolDefinition."),
    ("packages/ai-core", "Agent Contracts", "src/agents/Agent.ts", "Agent contract + BaseAgent abstract class."),
    ("packages/ai-core", "Prompts", "src/prompts/promptTemplate.ts", "PromptTemplate ({{var}} interpolation) + fillTemplate() helper."),
    ("packages/ai-core", "Barrel", "src/index.ts", "Re-exports all AI contracts."),

    # --- server: entry & config ---
    ("apps/server", "Entry", "src/index.ts", "Server bootstrap: listens on PORT and starts the scheduler."),
    ("apps/server", "Entry", "src/app.ts", "Express app setup: middlewares (helmet, cors), router mounting, error handler."),
    ("apps/server", "Config", "src/config/env.ts", "Validates environment variables with Zod (fails fast if missing)."),
    ("apps/server", "Config", "src/config/supabase.ts", "Initializes the Supabase client with the service-role key."),
    ("apps/server", "Config", "src/config/gemini.ts", "Initializes the Gemini SDK using GEMINI_API_KEY."),

    # --- server: AI / LLM ---
    ("apps/server", "AI · LLM", "src/ai/llm/geminiClient.ts", "Unified Gemini client (text/JSON/streaming/embeddings) — implements LlmClient from ai-core."),
    ("apps/server", "AI · LLM", "src/ai/llm/prompts/investmentReport.prompt.ts", "Prompt template for the narrative part of the investment report."),
    ("apps/server", "AI · LLM", "src/ai/llm/prompts/riskAssessment.prompt.ts", "Prompt template for structured (JSON) risk assessment."),
    ("apps/server", "AI · LLM", "src/ai/llm/prompts/qa.prompt.ts", "Prompt template for grounded RAG Q&A (context is the only source of truth)."),
    ("apps/server", "AI · LLM", "src/ai/llm/prompts/propertyComparison.prompt.ts", "Prompt template for comparing and ranking properties."),

    # --- server: AI / RAG ---
    ("apps/server", "AI · RAG", "src/ai/rag/embeddings.ts", "Generates embedding vectors via Gemini."),
    ("apps/server", "AI · RAG", "src/ai/rag/chunking.ts", "Splits long text into overlapping chunks to keep semantic coherence."),
    ("apps/server", "AI · RAG", "src/ai/rag/vectorStore.ts", "Vector store adapter over Supabase pgvector (insert + similarity search)."),
    ("apps/server", "AI · RAG", "src/ai/rag/retriever.ts", "Embeds a query and runs semantic search — implements Retriever (ragRetriever)."),
    ("apps/server", "AI · RAG", "src/ai/rag/ingestion.ts", "Ingestion pipeline: load -> chunk -> embed -> store (+ CLI `npm run ingest`)."),

    # --- server: AI / Agents ---
    ("apps/server", "AI · Agents", "src/ai/agents/orchestrator.ts", "Orchestrator: runs the full analysis pipeline (property + assumptions -> report)."),
    ("apps/server", "AI · Agents", "src/ai/agents/dataCollectorAgent.ts", "Collects and normalizes property data before analysis."),
    ("apps/server", "AI · Agents", "src/ai/agents/marketDataAgent.ts", "Collects market context (RAG + trends + neighborhood)."),
    ("apps/server", "AI · Agents", "src/ai/agents/calculationAgent.ts", "Computes financial metrics deterministically via shared-utils (never from the LLM)."),
    ("apps/server", "AI · Agents", "src/ai/agents/reportGeneratorAgent.ts", "Generates the narrative report + risk assessment via the LLM."),
    ("apps/server", "AI · Agents", "src/ai/agents/monitoringAgent.ts", "Monitors watched properties and raises alerts (called by the scheduled job)."),

    # --- server: HTTP ---
    ("apps/server", "HTTP · Routes", "src/routes/index.ts", "Root API router — mounts all feature routers under /api."),
    ("apps/server", "HTTP · Routes", "src/routes/auth.routes.ts", "Authentication routes (register/login)."),
    ("apps/server", "HTTP · Routes", "src/routes/property.routes.ts", "Property routes (search/detail)."),
    ("apps/server", "HTTP · Routes", "src/routes/analysis.routes.ts", "Analysis routes (/metrics, /compare)."),
    ("apps/server", "HTTP · Routes", "src/routes/report.routes.ts", "Report routes (generate/get/list) — auth protected."),
    ("apps/server", "HTTP · Routes", "src/routes/watchlist.routes.ts", "Watchlist routes."),
    ("apps/server", "HTTP · Routes", "src/routes/chat.routes.ts", "Chat route (RAG Q&A)."),
    ("apps/server", "HTTP · Controllers", "src/controllers/auth.controller.ts", "HTTP handler for auth: reads request, calls service, returns response."),
    ("apps/server", "HTTP · Controllers", "src/controllers/property.controller.ts", "HTTP handler for properties."),
    ("apps/server", "HTTP · Controllers", "src/controllers/analysis.controller.ts", "HTTP handler for analysis (computeMetrics, compare)."),
    ("apps/server", "HTTP · Controllers", "src/controllers/report.controller.ts", "HTTP handler for reports (generate, getById, list)."),
    ("apps/server", "HTTP · Controllers", "src/controllers/watchlist.controller.ts", "HTTP handler for the watchlist."),
    ("apps/server", "HTTP · Controllers", "src/controllers/chat.controller.ts", "HTTP handler for chat: retrieves context, builds prompt, returns answer + sources."),
    ("apps/server", "HTTP · Middleware", "src/middleware/auth.middleware.ts", "Verifies the token and injects req.user."),
    ("apps/server", "HTTP · Middleware", "src/middleware/error.middleware.ts", "Central error handler."),
    ("apps/server", "HTTP · Middleware", "src/middleware/logger.middleware.ts", "Request logging (pino-http)."),
    ("apps/server", "HTTP · Middleware", "src/middleware/rateLimit.middleware.ts", "Rate limiting (notably aiLimiter for costly AI routes)."),
    ("apps/server", "HTTP · Middleware", "src/middleware/validate.middleware.ts", "Validates request bodies with Zod schemas."),
    ("apps/server", "HTTP · Validators", "src/validators/auth.validator.ts", "Zod schemas for auth requests."),
    ("apps/server", "HTTP · Validators", "src/validators/property.validator.ts", "Zod schemas for property requests."),
    ("apps/server", "HTTP · Validators", "src/validators/analysis.validator.ts", "Zod schemas for analysis/report/chat requests."),

    # --- server: business & data ---
    ("apps/server", "Services", "src/services/auth.service.ts", "Authentication business logic."),
    ("apps/server", "Services", "src/services/property.service.ts", "Property business logic."),
    ("apps/server", "Services", "src/services/analysis.service.ts", "On-demand metric computation + LLM property comparison."),
    ("apps/server", "Services", "src/services/report.service.ts", "Runs the agent pipeline and persists the resulting report."),
    ("apps/server", "Services", "src/services/watchlist.service.ts", "Watchlist business logic."),
    ("apps/server", "Services", "src/services/market.service.ts", "Market data (trends/neighborhood)."),
    ("apps/server", "Services", "src/services/notification.service.ts", "Creates alerts/notifications."),
    ("apps/server", "Data · Repositories", "src/repositories/property.repository.ts", "Database access (Supabase) for properties + market."),
    ("apps/server", "Data · Repositories", "src/repositories/report.repository.ts", "Database access for reports."),
    ("apps/server", "Data · Repositories", "src/repositories/user.repository.ts", "Database access for users."),
    ("apps/server", "Data · Repositories", "src/repositories/watchlist.repository.ts", "Database access for the watchlist."),
    ("apps/server", "Data · Models", "src/models/property.model.ts", "Maps DB rows <-> Property domain type."),
    ("apps/server", "Data · Models", "src/models/report.model.ts", "Maps DB rows <-> InvestmentReport."),
    ("apps/server", "Data · Models", "src/models/user.model.ts", "Maps DB rows <-> UserProfile."),
    ("apps/server", "Data · Models", "src/models/watchlist.model.ts", "Maps DB rows <-> WatchlistItem."),
    ("apps/server", "Jobs", "src/jobs/scheduler.ts", "Schedules recurring jobs (node-cron)."),
    ("apps/server", "Jobs", "src/jobs/propertyMonitor.job.ts", "Recurring job that monitors watched properties."),
    ("apps/server", "Utils", "src/utils/apiError.ts", "Unified error class + helpers (badRequest/unauthorized/notFound)."),
    ("apps/server", "Utils", "src/utils/apiResponse.ts", "Unified response helpers (ok, created)."),
    ("apps/server", "Utils", "src/utils/asyncHandler.ts", "Wrapper that catches async errors in controllers."),
    ("apps/server", "Utils", "src/utils/financialCalc.ts", "Extra server-side financial helpers."),
    ("apps/server", "Utils", "src/utils/logger.ts", "Application logger (pino)."),
    ("apps/server", "Types", "src/types/index.ts", "Express type augmentation (e.g. req.user)."),

    # --- web ---
    ("apps/web", "Entry", "src/main.tsx", "Entry point — mounts the app into the DOM."),
    ("apps/web", "Entry", "src/App.tsx", "Root component + providers."),
    ("apps/web", "Routing", "src/routes/AppRouter.tsx", "Route definitions (react-router)."),
    ("apps/web", "Routing", "src/routes/routes.ts", "Page path constants."),
    ("apps/web", "Config", "src/config/env.ts", "Frontend env vars (API URL, Supabase keys)."),
    ("apps/web", "Services", "src/services/api/apiClient.ts", "Base HTTP client (axios) + interceptors."),
    ("apps/web", "Services", "src/services/api/authService.ts", "Auth endpoint calls."),
    ("apps/web", "Services", "src/services/api/propertyService.ts", "Property endpoint calls."),
    ("apps/web", "Services", "src/services/api/analysisService.ts", "Analysis endpoint calls."),
    ("apps/web", "Services", "src/services/api/reportService.ts", "Report endpoint calls."),
    ("apps/web", "Services", "src/services/api/watchlistService.ts", "Watchlist endpoint calls."),
    ("apps/web", "Services", "src/services/api/chatService.ts", "Chat endpoint calls."),
    ("apps/web", "Services", "src/services/supabase/supabaseClient.ts", "Frontend Supabase client (auth)."),
    ("apps/web", "State", "src/store/authStore.ts", "Global auth state (Zustand)."),
    ("apps/web", "State", "src/store/uiStore.ts", "UI state (modals, theme, toasts)."),
    ("apps/web", "ViewModels", "src/viewmodels/useAuthViewModel.ts", "Auth screen logic (state, calls, handlers)."),
    ("apps/web", "ViewModels", "src/viewmodels/useDashboardViewModel.ts", "Dashboard screen logic."),
    ("apps/web", "ViewModels", "src/viewmodels/usePropertySearchViewModel.ts", "Property search screen logic."),
    ("apps/web", "ViewModels", "src/viewmodels/usePropertyAnalysisViewModel.ts", "Property analysis screen logic."),
    ("apps/web", "ViewModels", "src/viewmodels/useReportViewModel.ts", "Report screen logic."),
    ("apps/web", "ViewModels", "src/viewmodels/useChatViewModel.ts", "Chat screen logic."),
    ("apps/web", "ViewModels", "src/viewmodels/useWatchlistViewModel.ts", "Watchlist screen logic."),
    ("apps/web", "Views · Components", "src/views/components/common/", "Generic UI: Button, Card, Input, Loader, Modal."),
    ("apps/web", "Views · Components", "src/views/components/layout/", "Layout: Navbar, Sidebar, Footer."),
    ("apps/web", "Views · Components", "src/views/components/property/", "Property UI: PropertyCard, PropertyList, PropertyFilters."),
    ("apps/web", "Views · Components", "src/views/components/analysis/", "Analysis charts: ROIChart, MarketTrendChart, RentalYieldCard, RiskMeter."),
    ("apps/web", "Views · Components", "src/views/components/report/ReportViewer.tsx", "Renders the investment report."),
    ("apps/web", "Views · Pages", "src/views/pages/", "Pages: Login, Register, Dashboard, PropertySearch, PropertyDetail, Analysis, Report, Chat, Watchlist."),
    ("apps/web", "Support", "src/models/, src/types/", "Web-specific models/types."),
    ("apps/web", "Support", "src/utils/", "constants, formatters, validators."),
    ("apps/web", "Support", "src/hooks/", "Generic hooks: useDebounce, useLocalStorage."),
    ("apps/web", "Support", "src/styles/theme.ts", "Theme/design tokens."),
    ("apps/web", "Config", "vite.config.ts", "Vite build config."),

    # --- mobile ---
    ("apps/mobile", "Entry", "App.tsx", "App root + providers."),
    ("apps/mobile", "Config", "app.json / babel.config.js", "Expo and Babel configuration."),
    ("apps/mobile", "Navigation", "src/navigation/RootNavigator.tsx", "Root navigator (switches between Auth and the app)."),
    ("apps/mobile", "Navigation", "src/navigation/AuthNavigator.tsx", "Auth screens stack."),
    ("apps/mobile", "Navigation", "src/navigation/TabNavigator.tsx", "Main app tabs."),
    ("apps/mobile", "Navigation", "src/navigation/types.ts", "Navigation param types."),
    ("apps/mobile", "Services", "src/services/api/, src/services/supabase/", "Same role as web (API calls + Supabase)."),
    ("apps/mobile", "State", "src/store/", "authStore, uiStore (Zustand)."),
    ("apps/mobile", "ViewModels", "src/viewmodels/", "Screen logic hooks (auth/dashboard/search/analysis/report/chat/watchlist)."),
    ("apps/mobile", "Views · Components", "src/views/components/", "common (Button/Card/Input/Loader), property, analysis (MetricCard/RiskMeter), report."),
    ("apps/mobile", "Views · Screens", "src/views/screens/", "Screens: Login, Register, Dashboard, PropertySearch, PropertyDetail, Analysis, Report, Chat, Watchlist."),
    ("apps/mobile", "Support", "src/theme/, src/hooks/, src/utils/, src/models/, src/types/, src/config/env.ts", "Theme, helpers, types and config."),

    # --- supabase ---
    ("supabase", "Database", "config.toml", "Local Supabase project config."),
    ("supabase", "Database", "migrations/0001_init.sql", "Base init (extensions, profiles)."),
    ("supabase", "Database", "migrations/0002_properties.sql", "Property tables."),
    ("supabase", "Database", "migrations/0003_reports.sql", "Investment report tables."),
    ("supabase", "Database", "migrations/0004_watchlist.sql", "Watchlist and alert tables."),
    ("supabase", "Database", "migrations/0005_rag.sql", "RAG tables (rag_documents, rag_chunks) + match_rag_chunks (pgvector)."),
    ("supabase", "Database", "seed.sql", "Seed data for testing."),

    # --- docs ---
    ("docs", "Docs", "ARCHITECTURE.md", "Overall project architecture."),
    ("docs", "Docs", "AI_DESIGN.md", "AI system design (agents + RAG + LLM)."),
    ("docs", "Docs", "API.md", "Endpoint documentation."),
    ("docs", "Docs", "DATA_MODEL.md", "Data model and relationships."),
    ("docs", "Docs", "FILE_STRUCTURE.md", "File-structure breakdown (Markdown source of this sheet)."),
]

HEADERS = ["Area", "Layer", "File", "What it does"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "File Structure"

# Styles
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
area_font = Font(bold=True, color="1F4E78")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(vertical="top", wrap_text=True)
header_align = Alignment(vertical="center", horizontal="left")

# Header row
ws.append(HEADERS)
for col in range(1, len(HEADERS) + 1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = header_align
    c.border = border

# Data rows with zebra striping + grouping color per area change
stripe = PatternFill("solid", fgColor="F2F6FC")
prev_area = None
for r, row in enumerate(ROWS, start=2):
    ws.append(row)
    is_new_area = row[0] != prev_area
    prev_area = row[0]
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=r, column=col)
        c.alignment = wrap_top
        c.border = border
        if r % 2 == 0:
            c.fill = stripe
    if is_new_area:
        ws.cell(row=r, column=1).font = area_font

# Column widths
widths = [20, 22, 46, 78]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header + enable autofilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

# --- Second sheet: Overview / golden rule ---
ws2 = wb.create_sheet("Overview")
overview = [
    ("PropertyPulse", "AI-Powered Real Estate Investment Advisor (monorepo)."),
    ("Structure", "apps/* = applications (web, mobile, server). packages/* = shared code."),
    ("shared-types", "Single source of truth for data shapes across all apps."),
    ("shared-utils", "Pure financial calculations and formatters."),
    ("ai-core", "Abstract AI contracts (LlmClient, Retriever, Agent) implemented by the server."),
    ("server flow", "Routes -> Controllers -> Services -> (Agents / Repositories) -> DB / LLM."),
    ("web & mobile flow", "Views -> ViewModels -> Services/API -> server (MVVM)."),
    ("Golden rule", "Financial numbers are computed deterministically in shared-utils, NOT by the LLM. The LLM only explains and recommends, keeping reports trustworthy and reproducible."),
]
ws2.append(["Topic", "Description"])
for col in range(1, 3):
    c = ws2.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.border = border
for r, row in enumerate(overview, start=2):
    ws2.append(row)
    ws2.cell(row=r, column=1).font = area_font
    for col in range(1, 3):
        cc = ws2.cell(row=r, column=col)
        cc.alignment = wrap_top
        cc.border = border
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 95
ws2.freeze_panes = "A2"

out = "docs/FILE_STRUCTURE.xlsx"
wb.save(out)
print(f"Wrote {out} with {len(ROWS)} rows across {len(wb.sheetnames)} sheets: {wb.sheetnames}")
